from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .dashboard import write_dashboard as write_dashboard

CATEGORY_SHEETS = ["Chicken", "Potatoes", "Vegetables", "Desserts", "Beef", "Pork", "Seafood", "Breakfast", "Snacks"]


def _df(rows) -> pd.DataFrame:
    if isinstance(rows, pd.DataFrame):
        return rows.copy()
    if isinstance(rows, dict):
        rows = [rows]
    return pd.DataFrame(list(rows or []))


def _style_workbook(path: str) -> None:
    workbook = load_workbook(path)
    for worksheet in workbook.worksheets:
        if worksheet.max_row >= 1:
            worksheet.freeze_panes = "A2"
            worksheet.auto_filter.ref = worksheet.dimensions
            for cell in worksheet[1]:
                cell.font = Font(bold=True)
                cell.fill = PatternFill("solid", fgColor="D9EAF7")
                cell.alignment = Alignment(vertical="center", wrap_text=True)
        for column_index, column in enumerate(worksheet.columns, 1):
            values = [str(cell.value) if cell.value is not None else "" for cell in column[:100]]
            width = min(48, max(10, max((len(value) for value in values), default=10) + 2))
            worksheet.column_dimensions[get_column_letter(column_index)].width = width
        for row in worksheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
    if "Rating Trends" in workbook.sheetnames:
        worksheet = workbook["Rating Trends"]
        if worksheet.max_row >= 3 and worksheet.max_column >= 2:
            chart = LineChart()
            chart.title = "Rating-count growth for current Top 10"
            chart.y_axis.title = "Rating count"
            chart.x_axis.title = "Observation timestamp"
            data = Reference(worksheet, min_col=2, max_col=worksheet.max_column, min_row=1, max_row=worksheet.max_row)
            categories = Reference(worksheet, min_col=1, min_row=2, max_row=worksheet.max_row)
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(categories)
            chart.height = 12
            chart.width = 24
            worksheet.add_chart(chart, "A18")
    workbook.save(path)


def write_csv_outputs(
    output_dir: str | Path,
    ranked: list[dict],
    coverage: list[dict],
    reliability: list[dict],
    anomalies: list[dict],
    *,
    source_health: list[dict] | None = None,
    robustness: list[dict] | None = None,
    dedupe_benchmark: list[dict] | None = None,
    pipeline_metrics: list[dict] | None = None,
    backtest: dict | None = None,
    evidence_calibration: list[dict] | None = None,
    evidence_label_results: list[dict] | None = None,
    quality_gate: dict | None = None,
    dedupe_label_queue: list[dict] | None = None,
) -> None:
    output = Path(output_dir)
    output.mkdir(parents=True, exist_ok=True)
    rankings = _df(ranked)
    if rankings.empty:
        (output / "leaderboard.csv").write_text("", encoding="utf-8")
        (output / "top50.csv").write_text("", encoding="utf-8")
    else:
        rankings.to_csv(output / "leaderboard.csv", index=False)
        rankings.head(50).to_csv(output / "top50.csv", index=False)
    _df(coverage).to_csv(output / "source_coverage.csv", index=False)
    _df(reliability).to_csv(output / "source_reliability.csv", index=False)
    _df(anomalies).to_csv(output / "anomalies.csv", index=False)
    _df(source_health).to_csv(output / "source_health.csv", index=False)
    _df(robustness).to_csv(output / "ranking_robustness.csv", index=False)
    _df(dedupe_benchmark).to_csv(output / "dedupe_benchmark.csv", index=False)
    _df(pipeline_metrics).to_csv(output / "pipeline_metrics.csv", index=False)
    backtest = backtest or {}
    _df(backtest.get("windows", [])).to_csv(output / "historical_backtest.csv", index=False)
    _df(backtest.get("configurations", [])).to_csv(output / "hyperparameter_evaluation.csv", index=False)
    _df(evidence_calibration).to_csv(output / "evidence_calibration.csv", index=False)
    _df(evidence_label_results).to_csv(output / "evidence_label_results.csv", index=False)
    _df(quality_gate).to_csv(output / "publication_quality_gate.csv", index=False)
    _df(dedupe_label_queue).to_csv(output / "dedupe_label_queue.csv", index=False)


def write_workbook(
    path: str | Path,
    ranked: list[dict],
    coverage: list[dict],
    reliability: list[dict],
    recent_observations: list[dict],
    anomalies: list[dict],
    duplicate_groups: list[dict],
    methodology: dict,
    *,
    source_health: list[dict] | None = None,
    uncertainty_calibration: list[dict] | None = None,
    robustness: list[dict] | None = None,
    dedupe_benchmark: list[dict] | None = None,
    pipeline_metrics: list[dict] | None = None,
    backtest: dict | None = None,
    evidence_calibration: list[dict] | None = None,
    evidence_label_results: list[dict] | None = None,
    quality_gate: dict | None = None,
    dedupe_label_queue: list[dict] | None = None,
    storage_health: dict | None = None,
    contracts: dict | None = None,
) -> None:
    path = str(path)
    rankings = _df(ranked)
    coverage_frame = _df(coverage)
    reliability_frame = _df(reliability)
    health_frame = _df(source_health)
    observations = _df(recent_observations)
    anomaly_frame = _df(anomalies)
    duplicates = _df(duplicate_groups)
    calibration = _df(uncertainty_calibration)
    robust = _df(robustness)
    benchmark = _df(dedupe_benchmark)
    metrics = _df(pipeline_metrics)
    evidence = _df(evidence_calibration)
    evidence_labels = _df(evidence_label_results)
    gate = _df(quality_gate)
    label_queue = _df(dedupe_label_queue)
    backtest = backtest or {}
    backtest_windows = _df(backtest.get("windows", []))
    hyperparameters = _df(backtest.get("configurations", []))
    backtest_summary = _df({key: value for key, value in backtest.items() if key not in {"windows", "configurations"}})
    storage = _df(storage_health)
    contract_frame = _df((contracts or {}).get("contracts", []))

    trend_chart = pd.DataFrame()
    if (
        not observations.empty
        and not rankings.empty
        and {"recipe_id", "timestamp", "rating_count"}.issubset(observations.columns)
    ):
        top_ids = [str(value) for value in rankings.head(10)["recipe_id"].tolist()] if "recipe_id" in rankings else []
        trend_source = observations[observations["recipe_id"].astype(str).isin(top_ids)].copy()
        if not trend_source.empty:
            labels = (
                dict(zip(rankings["recipe_id"].astype(str), rankings["title"].astype(str), strict=True))
                if {"recipe_id", "title"}.issubset(rankings.columns)
                else {}
            )
            trend_source["label"] = (
                trend_source["recipe_id"].astype(str).map(labels).fillna(trend_source["recipe_id"].astype(str))
            )
            trend_source["timestamp"] = pd.to_datetime(trend_source["timestamp"], errors="coerce", utc=True)
            trend_source = trend_source.dropna(subset=["timestamp"])
            trend_chart = (
                trend_source.pivot_table(index="timestamp", columns="label", values="rating_count", aggfunc="last")
                .sort_index()
                .reset_index()
            )
            if "timestamp" in trend_chart:
                trend_chart["timestamp"] = trend_chart["timestamp"].dt.tz_localize(None)

    movers = pd.DataFrame()
    entrants = pd.DataFrame()
    provenance = pd.DataFrame()
    time_signals = pd.DataFrame()
    if not rankings.empty:
        if "movement" in rankings:
            movers = rankings[rankings["movement"].notna()].copy()
            if not movers.empty:
                movers["abs_movement"] = movers["movement"].abs()
                movers = movers.sort_values(["abs_movement", "rating_count"], ascending=[False, False]).drop(
                    columns=["abs_movement"]
                )
        if "previous_rank" in rankings:
            entrants = rankings[rankings["previous_rank"].isna()].head(100).copy()
        provenance_columns = [
            column
            for column in (
                "rank",
                "title",
                "source",
                "rating",
                "rating_count",
                "category_expected_rating",
                "source_bias",
                "adjusted_rating",
                "posterior_mean",
                "uncertainty_penalty",
                "uncertainty_method",
                "evidence_penalty",
                "evidence_grade",
                "hierarchical_score",
                "rank_confidence",
                "rank_range_low",
                "rank_range_high",
                "rank_provenance",
                "url",
            )
            if column in rankings.columns
        ]
        provenance = rankings[provenance_columns].head(200).copy()
        time_columns = [
            column
            for column in (
                "rank",
                "title",
                "source",
                "review_growth_7d",
                "review_growth_30d",
                "rating_trend_30d",
                "rating_slope_30d_per_day",
                "review_slope_30d_per_day",
                "review_velocity_7d",
                "review_acceleration_14d",
                "page_change_count_30d",
                "last_material_page_change_at",
                "rating_change_point_30d",
                "rating_change_point_delta",
                "peak_rank",
                "rank_volatility",
                "url",
            )
            if column in rankings.columns
        ]
        time_signals = rankings[time_columns].head(500).copy()

    Path(path).parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        rankings.head(50).to_excel(writer, index=False, sheet_name="Top 50")
        rankings.to_excel(writer, index=False, sheet_name="All Rankings")
        provenance.to_excel(writer, index=False, sheet_name="Rank Explainability")
        coverage_frame.to_excel(writer, index=False, sheet_name="Source Coverage")
        health_frame.to_excel(writer, index=False, sheet_name="Source Health")
        reliability_frame.to_excel(writer, index=False, sheet_name="Source Reliability")
        observations.to_excel(writer, index=False, sheet_name="Rating History")
        trend_chart.to_excel(writer, index=False, sheet_name="Rating Trends")
        time_signals.to_excel(writer, index=False, sheet_name="Time Signals")
        calibration.to_excel(writer, index=False, sheet_name="Uncertainty Calibration")
        evidence.to_excel(writer, index=False, sheet_name="Evidence Calibration")
        evidence_labels.to_excel(writer, index=False, sheet_name="Evidence Labels")
        robust.to_excel(writer, index=False, sheet_name="Rank Robustness")
        backtest_summary.to_excel(writer, index=False, sheet_name="Backtest Summary")
        backtest_windows.to_excel(writer, index=False, sheet_name="Historical Backtest")
        hyperparameters.to_excel(writer, index=False, sheet_name="Hyperparameter Eval")
        metrics.to_excel(writer, index=False, sheet_name="Pipeline Metrics")
        gate.to_excel(writer, index=False, sheet_name="Publication Gate")
        storage.to_excel(writer, index=False, sheet_name="Storage Health")
        contract_frame.to_excel(writer, index=False, sheet_name="Data Contracts")
        entrants.to_excel(writer, index=False, sheet_name="New Entrants")
        movers.head(200).to_excel(writer, index=False, sheet_name="Biggest Movers")
        anomaly_frame.to_excel(writer, index=False, sheet_name="QA Anomalies")
        duplicates.to_excel(writer, index=False, sheet_name="Duplicate Groups")
        benchmark.to_excel(writer, index=False, sheet_name="Dedupe Benchmark")
        label_queue.to_excel(writer, index=False, sheet_name="Dedupe Label Queue")
        pd.DataFrame([methodology]).to_excel(writer, index=False, sheet_name="Methodology")
        for category in CATEGORY_SHEETS:
            if rankings.empty or "categories" not in rankings:
                subset = pd.DataFrame()
            else:
                subset = rankings[rankings["categories"].fillna("").str.contains(category, regex=False)].head(100)
            subset.to_excel(writer, index=False, sheet_name=category)
    _style_workbook(path)
