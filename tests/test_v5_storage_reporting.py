from __future__ import annotations

import json
from pathlib import Path

from openpyxl import load_workbook

from airfryer_rankings.evidence_calibration import apply_evidence_calibration, evaluate_evidence_labels
from airfryer_rankings.model_config import ModelParameters, load_model_config
from airfryer_rankings.models import RecipeRow
from airfryer_rankings.quality_gate import assert_publishable, evaluate_publish_gate, load_previous_serving_snapshot
from airfryer_rankings.reporting import write_csv_outputs, write_workbook
from airfryer_rankings.storage import load_state, merge_observations, read_recent_records, save_state, write_run_records


def _row(confidence: float = 0.65) -> RecipeRow:
    return RecipeRow(
        recipe_id="fixture",
        title="Air Fryer Fixture",
        source="example.com",
        url="https://example.com/fixture",
        rating=4.8,
        rating_count=100,
        best_rating=5.0,
        normalized_rating=4.8,
        retrieved_at="2026-08-18T20:00:00+00:00",
        canonical_url="https://example.com/fixture",
        evidence_confidence=confidence,
        evidence_status="schema_only",
        page_hash="page",
        categories=("Chicken",),
    )


def test_storage_round_trip_migration_and_recent_records(tmp_path: Path):
    state_path = tmp_path / "state.json"
    state_path.write_text(
        json.dumps(
            {
                "schema_version": 3,
                "recipes": {
                    "legacy": {
                        "recipe_id": "legacy",
                        "title": "Legacy",
                        "source": "example.com",
                        "url": "https://example.com/legacy",
                        "canonical_url": "https://example.com/legacy",
                        "normalized_rating": 4.9,
                        "rating_count": 500,
                        "evidence_confidence": 0.85,
                        "evidence_status": "",
                    }
                },
                "url_catalog": {},
            }
        ),
        encoding="utf-8",
    )
    state = load_state(state_path)
    assert state["schema_version"] == 4
    assert state["recipes"]["legacy"]["clean_schema_version"] == 5
    assert state["recipes"]["legacy"]["needs_evidence_backfill"] is True

    observations = merge_observations(state, [_row()], "2026-08-18T20:00:00+00:00")
    assert observations[0]["schema_version"] == 2
    save_state(state_path, state)
    reloaded = load_state(state_path)
    assert reloaded["recipes"]["fixture"]["clean_schema_version"] == 5

    history = tmp_path / "history"
    first = write_run_records(history, observations, "2026-08-18T20:00:00+00:00")
    assert first is not None
    Path(first).write_text(Path(first).read_text(encoding="utf-8") + "not-json\n", encoding="utf-8")
    recent = read_recent_records(history, limit=10)
    assert recent and recent[-1]["recipe_id"] == "fixture"
    assert write_run_records(history, [], "2026-08-18T21:00:00+00:00") is None


def test_model_configuration_loading_and_frozen_overrides(tmp_path: Path):
    config = tmp_path / "model.yaml"
    config.write_text(
        """model_version: 7
active:
  max_source_bias: 0.12
  evidence_confidence_target: 0.82
  evidence_penalty_scale: 0.18
  uncertainty_cap: 0.22
  source_prior_strength: 25
  category_prior_strength: 30
  volume_prior_quantile: 0.65
  minimum_volume_prior: 75
  volume_prior_multiplier: 1.1
""",
        encoding="utf-8",
    )
    params, payload = load_model_config(config)
    assert payload["model_version"] == 7
    assert params.max_source_bias == 0.12
    assert params.minimum_volume_prior == 75
    changed = params.with_overrides(max_source_bias=0.2)
    assert changed.max_source_bias == 0.2
    assert params.max_source_bias == 0.12
    assert isinstance(ModelParameters(), ModelParameters)


def test_evidence_calibration_activates_only_after_review_threshold(tmp_path: Path):
    fixture_root = tmp_path / "fixtures"
    fixture_root.mkdir()
    fixture = fixture_root / "schema.html"
    fixture.write_text(
        '<html><head><script type="application/ld+json">'
        '{"@type":"Recipe","name":"Air Fryer X","recipeIngredient":["chicken"],'
        '"aggregateRating":{"ratingValue":"4.8","ratingCount":"100","bestRating":"5"}}'
        "</script></head><body></body></html>",
        encoding="utf-8",
    )
    labels = {
        "labels": [
            {
                "id": f"label-{index}",
                "fixture": "schema.html",
                "source": "example.com",
                "url": "https://example.com/x",
                "expected_status": "schema_only",
                "expected_rating": 4.8,
                "expected_rating_count": 100,
            }
            for index in range(30)
        ]
    }
    label_path = tmp_path / "labels.json"
    label_path.write_text(json.dumps(labels), encoding="utf-8")
    calibration, outcomes = evaluate_evidence_labels(label_path, fixture_root=fixture_root)
    assert len(outcomes) == 30
    assert calibration["schema_only"]["ready"] is True
    assert calibration["schema_only"]["empirical_accuracy"] == 1.0
    calibrated = apply_evidence_calibration([_row(0.65)], calibration)
    assert calibrated[0].evidence_confidence == 1.0
    missing, missing_rows = evaluate_evidence_labels(tmp_path / "missing.json", fixture_root=fixture_root)
    assert missing == {} and missing_rows == []


def test_reporting_writes_v5_csvs_and_workbook(tmp_path: Path):
    ranked = [
        {
            "rank": 1,
            "recipe_id": "a",
            "title": "Air Fryer Chicken",
            "source": "example.com",
            "url": "https://example.com/a",
            "rating": 4.9,
            "rating_count": 1000,
            "hierarchical_score": 4.7,
            "evidence_confidence": 1.0,
            "evidence_status": "verified",
            "evidence_grade": "A+",
            "rank_confidence": 0.95,
            "rank_range_low": 1,
            "rank_range_high": 2,
            "rank_provenance": "test",
            "categories": "Chicken",
            "movement": 1,
            "previous_rank": 2,
            "review_growth_7d": 10,
            "review_growth_30d": 30,
            "rating_trend_30d": 0.01,
            "rating_slope_30d_per_day": 0.001,
            "review_slope_30d_per_day": 2.0,
            "review_velocity_7d": 2.5,
            "review_acceleration_14d": 0.5,
            "peak_rank": 1,
            "rank_volatility": 0.5,
        }
    ]
    observations = [{"recipe_id": "a", "timestamp": "2026-08-18T20:00:00+00:00", "rating_count": 1000, "rating": 4.9}]
    output = tmp_path / "output"
    write_csv_outputs(
        output,
        ranked,
        [{"source": "example.com", "status": "ok"}],
        [{"source": "example.com", "run_success_rate": 1.0}],
        [],
        source_health=[{"source": "example.com", "checked_this_run": True}],
        robustness=[{"max_source_bias": 0.15}],
        dedupe_benchmark=[{"pair_id": "x", "outcome": "TP"}],
        pipeline_metrics=[{"metric": "ranked_recipes", "value": 1}],
        backtest={"windows": [{"recipes": 1}], "configurations": [{"config_id": "active"}]},
        evidence_calibration=[{"evidence_status": "verified", "ready": False}],
        evidence_label_results=[{"label_id": "x", "correct": True}],
        quality_gate={"passed": True},
        dedupe_label_queue=[{"pair_id": "candidate"}],
    )
    assert (output / "leaderboard.csv").exists()
    assert (output / "historical_backtest.csv").exists()

    workbook_path = output / "rankings.xlsx"
    write_workbook(
        workbook_path,
        ranked,
        [{"source": "example.com", "status": "ok"}],
        [{"source": "example.com", "run_success_rate": 1.0}],
        observations,
        [],
        [],
        {"model_version": 5},
        source_health=[{"source": "example.com", "checked_this_run": True}],
        uncertainty_calibration=[{"bucket": "100-499", "ready": False}],
        robustness=[{"max_source_bias": 0.15}],
        dedupe_benchmark=[{"pair_id": "x", "outcome": "TP"}],
        pipeline_metrics=[{"metric": "ranked_recipes", "value": 1}],
        backtest={"ready": False, "windows": [], "configurations": []},
        evidence_calibration=[{"evidence_status": "verified", "ready": False}],
        evidence_label_results=[{"label_id": "x", "correct": True}],
        quality_gate={"passed": True},
        dedupe_label_queue=[{"pair_id": "candidate"}],
        storage_health={"archive_recommended": False},
        contracts={"contracts": [{"name": "raw_observation", "version": 2}]},
    )
    workbook = load_workbook(workbook_path, read_only=True)
    expected_sheets = {
        "Top 50",
        "Rank Explainability",
        "Time Signals",
        "Evidence Calibration",
        "Backtest Summary",
        "Pipeline Metrics",
        "Publication Gate",
        "Storage Health",
        "Data Contracts",
        "Dedupe Label Queue",
    }
    assert expected_sheets.issubset(set(workbook.sheetnames))


def test_publication_gate_pass_and_previous_snapshot_loading(tmp_path: Path):
    output = tmp_path / "output"
    output.mkdir()
    (output / "summary.json").write_text(json.dumps({"ranked_recipes": 50, "model_version": 5}), encoding="utf-8")
    (output / "leaderboard.csv").write_text("recipe_id,rank\na,1\n", encoding="utf-8")
    summary, previous = load_previous_serving_snapshot(output)
    assert summary["ranked_recipes"] == 50
    assert previous[0]["recipe_id"] == "a"

    ranked = [{"recipe_id": f"r{index}", "rank": index + 1} for index in range(50)]
    result = evaluate_publish_gate(
        {"ranked_recipes": 50, "model_version": 5, "deduplicated_count": 1},
        ranked,
        ranked,
        {"evidence_conflict_rate": 0.0, "legacy_evidence_pending": 0, "http_429": 0},
        mode="hourly",
        model_version=5,
        deduplicated_count=1,
    )
    assert result["passed"] is True
    assert_publishable(result)
