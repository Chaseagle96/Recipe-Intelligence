from openpyxl import load_workbook

from airfryer_rankings.reporting import write_dashboard, write_workbook


def sample_ranked():
    return [
        {
            "rank": 1,
            "recipe_id": "a",
            "title": "Air Fryer Chicken Bites",
            "source": "x.com",
            "combined_sources": "x.com",
            "url": "https://x.com/a",
            "rating": 4.9,
            "rating_count": 500,
            "source_bias": 0.02,
            "category_expected_rating": 4.75,
            "adjusted_rating": 4.88,
            "posterior_mean": 4.85,
            "uncertainty_penalty": 0.1,
            "uncertainty_method": "theoretical_max_variance",
            "evidence_penalty": 0.0,
            "hierarchical_score": 4.75,
            "evidence_confidence": 1.0,
            "evidence_grade": "A",
            "evidence_status": "verified",
            "author": "A",
            "categories": "Chicken | Snacks",
            "duplicate_group_id": "",
            "duplicate_confidence": 0.0,
            "last_seen_at": "2026-08-18T20:00:00+00:00",
            "rating_change": 0.0,
            "review_count_change": 2,
            "review_velocity_per_day": 48.0,
            "review_growth_7d": 20,
            "review_growth_30d": 60,
            "rating_trend_30d": 0.02,
            "peak_rank": 1,
            "rank_volatility": 0.5,
            "days_in_top10": 3,
            "days_in_top50": 3,
            "rank_confidence": 0.94,
            "rank_stddev": 0.4,
            "rank_range_low": 1,
            "rank_range_high": 3,
            "top10_frequency": 1.0,
            "top50_frequency": 1.0,
            "previous_rank": 2,
            "movement": 1,
            "rank_provenance": "Raw 4.900; source/category adjustment -0.020; posterior 4.850; uncertainty -0.100; evidence -0.000; final 4.750.",
        }
    ]


def test_workbook_contains_research_sheets(tmp_path):
    path = tmp_path / "rankings.xlsx"
    write_workbook(
        path,
        sample_ranked(),
        [],
        [],
        [],
        [],
        [],
        {"formula": "test"},
        source_health=[],
        uncertainty_calibration=[],
        robustness=[],
        dedupe_benchmark=[],
    )
    wb = load_workbook(path, read_only=True)
    expected = {
        "Top 50",
        "All Rankings",
        "Rank Explainability",
        "Source Coverage",
        "Source Health",
        "Source Reliability",
        "Rating History",
        "Rating Trends",
        "Uncertainty Calibration",
        "Rank Robustness",
        "New Entrants",
        "Biggest Movers",
        "QA Anomalies",
        "Duplicate Groups",
        "Dedupe Benchmark",
        "Methodology",
        "Chicken",
        "Potatoes",
        "Vegetables",
        "Desserts",
        "Beef",
        "Pork",
        "Seafood",
        "Breakfast",
        "Snacks",
    }
    assert expected.issubset(set(wb.sheetnames))


def test_dashboard_is_searchable_and_exposes_confidence(tmp_path):
    write_dashboard(tmp_path, "2026-08-18T20:00:00+00:00", sample_ranked(), [], [], {"formula": "test"}, 40)
    html = (tmp_path / "index.html").read_text()
    data = (tmp_path / "data.json").read_text()
    assert 'id="search"' in html
    assert 'id="rankconfidence"' in html
    assert "Why this rank?" in html
    assert "Air Fryer Chicken Bites" in data
    assert '"rank_confidence": 0.94' in data


def test_rating_trends_sheet_contains_growth_chart(tmp_path):
    path = tmp_path / "rankings.xlsx"
    observations = [
        {"recipe_id": "a", "timestamp": "2026-08-18T19:00:00+00:00", "rating_count": 490},
        {"recipe_id": "a", "timestamp": "2026-08-18T20:00:00+00:00", "rating_count": 500},
    ]
    write_workbook(path, sample_ranked(), [], [], observations, [], [], {"formula": "test"})
    wb = load_workbook(path)
    assert len(wb["Rating Trends"]._charts) == 1
